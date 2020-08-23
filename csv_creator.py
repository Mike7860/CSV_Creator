import os
import openpyxl
import PyPDF2
import csv
import textwrap
import logging
import time
import datetime

from datetime import date
from openpyxl.workbook import Workbook

curr_dir = os.getcwd()

print('You are in "' +curr_dir+ '" Do you want change directory? y/n')

dir_choise = input()

if dir_choise =='y':
    print('Please provide new directory with necessary files \n>>')
    os.chdir(input())
if dir_choise =='n':
    curr_dir = curr_dir

not_tested = 0
date = date.today()
print(date)

#main class
class Test:
    def __init__(self, id, source, flag, auto):
        self.ID = id
        self.source = source
        self.flag = flag
        self.auto = auto

#dictionaries declaration
dict_test_all = {}
dict_test_basic = {}
dict_test_extended = {}

with open('all.csv', newline='') as csvfile:
    all = csv.reader(csvfile)
    for row in all:
        #
        if row[0] == "Test ID" and row[1] == "Source" and row[2] == "Flag" and row[3] == "Automate":
            continue
        dict_test_all[row[0]] = Test(row[0],row[1],row[2],row[3])
workbook_1 = openpyxl.load_workbook('not_test_1.xlsx', 'rb', data_only=True)
not_test_1 =  workbook_1['not_tested']
all_rows = not_test_1.max_row - 1
nt1 = "D" + str(all_rows+1)
nt1_range = not_test_1['A2':nt1]

print('Please select scope of testing: \n1) basic \n2) extended\n>>')
choise = input ()

if choise == '1':
    for A2, B2, C2, D2 in nt1_range:
        if str(B2.value) != "0" and dict_test_all[B2.value].ID == str(B2.value):
            dict_test_all[B2.value].flag = "2"
            print("Not_testable: %s" %(dict_test_all[B2.value].ID))
            not_tested = not_tested + 1
    print("Not tested (based on first file)", not_tested)

if choise == '2':
    for A2, B2, C2, D2 in nt1_range:
        if str(C2.value) != "0" and dict_test_all[C2.value].ID == str(C2.value):
            dict_test_all[C2.value].flag = "2"
            print("Not_testable: %s" %(dict_test_all[C2.value].ID))
            not_tested = not_tested + 1
    print("Not tested (based on first file)", not_tested)

workbook_2 = openpyxl.load_workbook('not_test_2.xlsx', 'rb')
tests =  workbook_2['tests']
scope = workbook_2['scope']

all_rows_2 = tests.max_row
nt2 = "D" + str(all_rows_2)

nt2_range = tests['A2':nt2]

for A2, B2, C2, D2 in nt2_range:
    if str(str(D2.value).lower()) == "false" and dict_test_all[A2.value].ID == str(A2.value):
        dict_test_all[A2.value].flag = "2"
        print("Not_testable: %s" % (dict_test_all[A2.value].ID))
        not_tested = not_tested + 1
        continue

all_rows_3 = scope.max_row
nt3 = "D" + str(all_rows_3)

nt3_range = scope['A2':nt3]
for A2, B2, C2, D2 in nt3_range:
    if str(str(D2.value)) == "basic" and dict_test_all[A2.value].ID == str(A2.value):
        dict_test_basic[A2.value] = Test(dict_test_all[A2.value].ID, dict_test_all[A2.value].source, dict_test_all[A2.value].flag, dict_test_all[A2.value].auto)
    if str(str(D2.value)) == "extended" and dict_test_all[A2.value].ID == str(A2.value):
        dict_test_extended[A2.value] = Test(dict_test_all[A2.value].ID, dict_test_all[A2.value].source, dict_test_all[A2.value].flag, dict_test_all[A2.value].auto)

if choise == '1':
    with open('final.csv', 'w') as csvfile:
        fieldnames = ["Test ID", "Source", "Flag", "Automate"]
        csv_write = csv.DictWriter(csvfile, fieldnames = fieldnames, delimiter = ',', lineterminator = '\n')
        csv_write.writeheader()

        for value in dict_test_basic.values():
            csv_write.writerow({
                "Test ID": value.ID,
                "Source": value.source,
                "Flag": value.flag,
                "Automate": value.auto
            })

    for key, value in dict_test_basic.items():
        Test_row = "key: %s | value: %s | %s | %s | %s" %(key, value.ID, value.source, value.flag, value.auto)
        logging.basicConfig(filename=str(date)+" logs.log", filemode='w', level=logging.INFO)
        logging.info(Test_row)

    print("All not tested: ", not_tested)
    print("All tests: ", len(dict_test_basic))

if choise == '2':
    with open('final_test_plan.csv', 'w') as csvfile:
        fieldnames = ["Test ID", "Source", "Flag", "Automate"]
        csv_write = csv.DictWriter(csvfile, fieldnames = fieldnames, delimiter = ',', lineterminator = '\n')
        csv_write.writeheader()

        for value in dict_test_extended.values():
            csv_write.writerow({
                "Test ID": value.ID,
                "Source": value.source,
                "Flag": value.flag,
                "Automate": value.auto
            })

    for key, value in dict_test_extended.items():
        Test_row = "key: %s | value: %s | %s | %s | %s" %(key, value.ID, value.source, value.flag, value.auto)
        logging.basicConfig(filename=str(date)+" logs.log", filemode='w', level=logging.INFO)
        logging.info(Test_row)

    print("All not tested: ", not_tested)
    print("All tests: ", len(dict_test_extended))